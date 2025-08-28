from openpyxl import load_workbook, Workbook

file_path = "Employee Sample Data.xlsx"
try:
    wb = load_workbook(file_path)
    ws = wb.active  
except FileNotFoundError:
    print(f"Error: The file '{file_path}' was not found. Please ensure the file is in the same directory.")
    exit()

headers = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), start=1)}

if "Department" not in headers or "Annual Salary" not in headers:
    print("Error: The Excel file must contain 'Department' and 'Annual Salary' columns.")
    exit()

dept_col = headers["Department"]
salary_col = headers["Annual Salary"]

dept_data = {}

for row in ws.iter_rows(min_row=2, values_only=True):  
    dept = row[dept_col - 1]
    salary = row[salary_col - 1]

    if dept and isinstance(salary, (int, float)): 
        if dept not in dept_data:
            dept_data[dept] = {"total_salary": 0, "count": 0}
        dept_data[dept]["total_salary"] += salary
        dept_data[dept]["count"] += 1

summary_wb = Workbook()
summary_ws = summary_wb.active
summary_ws.title = "Summary"

summary_ws.append(["Department", "Avg_Salary", "Employee_Count"])

for dept, data in dept_data.items():
    if data["count"] > 0:  
        avg_salary = data["total_salary"] / data["count"]
        summary_ws.append([dept, round(avg_salary, 2), data["count"]])

try:
    summary_wb.save("Employee_Summary.xlsx")
    print("Automation complete ✅ Created: Employee_Summary.xlsx")
except Exception as e:
    print(f"Error: Could not save the file. {e}")
